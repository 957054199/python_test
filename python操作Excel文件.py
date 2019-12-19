# 创建一个工作簿

from openpyxl import Workbook

wb = Workbook()

# 新建一张表

wb.create_sheet("sheet0")

# 写入值
sheet = wb["sheet0"]
# 修改数据
sheet.cell(4, 1).value = "在天愿作比翼鸟"
# 增加数据
sheet.cell(4, 2, "在地愿做连理枝")
wb.save("python.xlsx")

# 操作excel load_workbook

from openpyxl import load_workbook

wb = load_workbook("python操作.xlsx")

# 定位表单
# wb.get_sheet_by_name("json")
sheet = wb["json"]

# 定位单元格cell 参数必须都是数字，参数是坐标
value = sheet.cell(1, 2).value
print(value)
A3 = sheet.cell(2, 1).value
print("A3：", A3)
print("A3的值是：%s，A3的类型是：%s" % (A3, type(A3)))
B3 = sheet.cell(2, 2).value
print("B3：", B3)
print("B3的值是：%s，B3的类型是：%s" % (B3, type(B3)))
C3 = sheet.cell(2, 3).value
print("C3：", C3)
print("C3的值是：%s，C3的类型是：%s" % (C3, type(C3)))
D3 = sheet.cell(3, 2).value
print("D3：", D3)
print("D3的值是：%s，D3的类型是：%s" % (D3, type(D3)))
E3 = sheet.cell(3, 3).value
print("E3：", E3)
print("E3的值是：%s，E3的类型是：%s" % (E3, type(E3)))

# eval()方法 把数据转换成python原本可以识别的数据类型
value = sheet.cell(1, 2).value
print(value)
A3 = sheet.cell(2, 1).value
print("A3：", A3)
print("A3的值是：%s，A3的类型是：%s" % (A3, type(A3)))
B3 = eval(sheet.cell(2, 2).value)
print("B3：", B3)
print("B3的值是：%s，B3的类型是：%s" % (B3, type(B3)))
C3 = eval(sheet.cell(2, 3).value)
print("C3：", C3)
print("C3的值是：%s，C3的类型是：%s" % (C3, type(C3)))
D3 = eval(sheet.cell(3, 2).value)
print("D3：", D3)
print("D3的值是：%s，D3的类型是：%s" % (D3, type(D3)))
E3 = eval(sheet.cell(3, 3).value)
print("E3：", E3)
print("E3的值是：%s，E3的类型是：%s" % (E3, type(E3)))

wb.save("python操作.xlsx")
